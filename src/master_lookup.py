r"""Master mapping lookup table — pre-flight cache for the pipeline.

Loaded by standardize_file.py at runtime to resolve known raw values without
an API call. Built once from the master remapping Excel file by running
src/build_lookup.py (or whenever the master file changes).

Lookup priority per value:
  1. Country-specific entry  (country, raw_value) — for keep_by_country fields
  2. Country-agnostic entry  (raw_value)          — for consistent/resolved entries
  3. None                                          — not in lookup → Claude API

Usage:
    # Build (one-time, run from project root)
    .venv\Scripts\python.exe src\build_lookup.py

    # In code
    from master_lookup import load_lookup
    lookup = load_lookup()
    result = lookup["positions_designations"].get("CEO", country="United Kingdom")
    # -> "Board Member"
"""

from __future__ import annotations

import json
import re
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field as dc_field
from difflib import SequenceMatcher
from pathlib import Path
from typing import Optional

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

import openpyxl

import config
import fields as field_registry

DEFAULT_MASTER_FILE = (
    Path.home() / "Downloads" / "[FINAL]- ALL Countries GG Standardization Remapping.xlsx"
)
DEFAULT_LOOKUP_FILE = config.DATA_DIR / "master_lookup.json"

SIMILARITY_EXAMPLE_MIN_SCORE = 0.45
SIMILARITY_AUTO_SCORE = 0.97
SIMILARITY_AUTO_MARGIN = 0.12


@dataclass(frozen=True)
class SimilarityMatch:
    """One approved lookup entry similar to a previously unseen raw value."""

    raw_value: str
    standardized_value: str
    score: float
    country: str = ""


@dataclass(frozen=True)
class SimilarityPrediction:
    """A conservative prediction backed by strongly agreeing lookup entries."""

    standardized_value: str
    score: float
    matches: tuple[SimilarityMatch, ...]

# Sheet names in the master file — hand-mapped (brief Section 9).
SHEET_BY_FIELD: dict[str, str] = {
    "business_legal_form":           "StandardizedIncorporationDetail",
    "positions_designations":        "Universal Position + Designatio",
    "business_status":               "Sheet20",
    "directors_officers_status":     "Status - DirectorsOfficers",
    "ownership_relationship_status": "Status - OwnershipRelationship",
    "psc_beneficiary_type":          "UniversalBeneficiaryType",
    "directors_officers_type":       "Type - DirectorsOfficers",
    "ownership_relationship_type":   "Type - Ownership Relationship T",
    "business_entity_type":          "BusinessEntityType",
    "brn_type":                      "BRN Type",
}

# Normalise legacy labels in the master file to the current taxonomy label.
_LABEL_FIX: dict[str, str] = {
    "Other / Unknown":  "Other / Unclassified",
    "Unknown":          "Other / Unclassified",
    "Owner/Controller": "Owner / Controller",   # spacing inconsistency in master file
}

# ---------------------------------------------------------------------------
# Conflict resolutions — authoritative source used by both apply_resolutions.py
# and build_lookup.py.
# Values: "keep_by_country" or a specific standardized-value string.
# ---------------------------------------------------------------------------

RESOLUTIONS: dict[tuple[str, str], str] = {

    # ------------------------------------------------------------------
    # POSITIONS / DESIGNATIONS (59) — all genuine country rules.
    # Within-country ambiguities are skipped automatically at build time.
    # ------------------------------------------------------------------
    ("positions_designations", "Administrator"):              "keep_by_country",
    ("positions_designations", "Assistant General Manager"):  "keep_by_country",
    ("positions_designations", "Associate"):                  "keep_by_country",
    ("positions_designations", "Commissioner"):               "keep_by_country",
    ("positions_designations", "Communication Director"):     "keep_by_country",
    ("positions_designations", "Controller"):                 "keep_by_country",
    ("positions_designations", "Director"):                   "keep_by_country",
    ("positions_designations", "Liquidating Agent"):          "keep_by_country",
    ("positions_designations", "Manager"):                    "keep_by_country",
    ("positions_designations", "Member"):                     "keep_by_country",
    ("positions_designations", "Officer"):                    "keep_by_country",
    ("positions_designations", "President"):                  "keep_by_country",
    ("positions_designations", "President,President"):        "keep_by_country",
    ("positions_designations", "Principal"):                  "keep_by_country",
    ("positions_designations", "Secretary"):                  "keep_by_country",
    ("positions_designations", "Sole Administrator"):         "keep_by_country",
    ("positions_designations", "Treasurer"):                  "keep_by_country",
    ("positions_designations", "administration director"):    "keep_by_country",
    ("positions_designations", "administrator"):              "keep_by_country",
    ("positions_designations", "administrator,partner"):      "keep_by_country",
    ("positions_designations", "ambassador"):                 "keep_by_country",
    ("positions_designations", "assistant director"):         "keep_by_country",
    ("positions_designations", "ceo"):                        "keep_by_country",
    ("positions_designations", "chief executive officer"):    "keep_by_country",
    ("positions_designations", "commissioner"):               "keep_by_country",
    ("positions_designations", "company administrator"):      "keep_by_country",
    ("positions_designations", "company formation agent"):    "keep_by_country",
    ("positions_designations", "controller"):                 "keep_by_country",
    ("positions_designations", "cotas em tesouraria"):        "keep_by_country",
    ("positions_designations", "director"):                   "keep_by_country",
    ("positions_designations", "director of finance"):        "keep_by_country",
    ("positions_designations", "director of operations"):     "keep_by_country",
    ("positions_designations", "entrepreneur"):               "keep_by_country",
    ("positions_designations", "executive"):                  "keep_by_country",
    ("positions_designations", "executive director"):         "keep_by_country",
    ("positions_designations", "financial director"):         "keep_by_country",
    ("positions_designations", "managing director"):          "keep_by_country",
    ("positions_designations", "managing director,director"): "keep_by_country",
    ("positions_designations", "managing director,partner"):  "keep_by_country",
    ("positions_designations", "marketing director"):         "keep_by_country",
    ("positions_designations", "mayor"):                      "keep_by_country",
    ("positions_designations", "operations director"):        "keep_by_country",
    ("positions_designations", "personnel director"):         "keep_by_country",
    ("positions_designations", "president"):                  "keep_by_country",
    ("positions_designations", "president,director"):         "keep_by_country",
    ("positions_designations", "presidente"):                 "keep_by_country",
    ("positions_designations", "principal"):                  "keep_by_country",
    ("positions_designations", "production director"):        "keep_by_country",
    ("positions_designations", "purchasing director"):        "keep_by_country",
    ("positions_designations", "sales director"):             "keep_by_country",
    ("positions_designations", "senior managing director"):   "keep_by_country",
    ("positions_designations", "socio-administrador"):        "keep_by_country",
    ("positions_designations", "sole managing administrator"): "keep_by_country",
    ("positions_designations", "sole trader"):                "keep_by_country",
    ("positions_designations", "superintendent"):             "keep_by_country",
    ("positions_designations", "supervisor"):                 "keep_by_country",
    ("positions_designations", "technical director"):         "keep_by_country",
    ("positions_designations", "trustee"):                    "keep_by_country",
    ("positions_designations", "tutor"):                      "keep_by_country",

    # ------------------------------------------------------------------
    # BUSINESS LEGAL FORM (6)
    # ------------------------------------------------------------------
    ("business_legal_form", "As"):            "keep_by_country",
    ("business_legal_form", "business name"): "keep_by_country",
    ("business_legal_form", "company"):       "Company",
    ("business_legal_form", "headquarters"):  "keep_by_country",
    ("business_legal_form", "non-stock"):     "Non-Profit / Cooperative",
    ("business_legal_form", "subsidiary"):    "keep_by_country",

    # ------------------------------------------------------------------
    # BUSINESS STATUS (1)
    # ------------------------------------------------------------------
    ("business_status", "Pending"):           "Pending / Insolvency",

    # ------------------------------------------------------------------
    # OWNERSHIPRELATIONSHIP STATUS (6)
    # ------------------------------------------------------------------
    ("ownership_relationship_status", "Active (Revived)"):             "Active",
    ("ownership_relationship_status", "Dissolved"):                     "Inactive",
    ("ownership_relationship_status", "Inactive - Cancellation of GP"): "Inactive",
    ("ownership_relationship_status", "Pending"):                       "Pending / Insolvency",
    ("ownership_relationship_status", "Resigned"):                      "Inactive",
    ("ownership_relationship_status", "Start"):                         "Other / Unclassified",

    # ------------------------------------------------------------------
    # PSC / BENEFICIARY TYPE (4)
    # ------------------------------------------------------------------
    ("psc_beneficiary_type", "ORDINARY"):                      "Owner / Beneficial Owner",
    ("psc_beneficiary_type", "Partner"):                       "keep_by_country",
    ("psc_beneficiary_type", "Person with Significant Control"): "Controller",
    ("psc_beneficiary_type", "unidentified share ownership"):   "Other / Unclassified",

    # ------------------------------------------------------------------
    # OWNERSHIPRELATIONSHIP TYPE (2)
    # ------------------------------------------------------------------
    ("ownership_relationship_type", "Share Ownership"):           "keep_by_country",
    ("ownership_relationship_type", "Unidentified Share Ownership"): "Other / Unclassified",

    # ------------------------------------------------------------------
    # BRN TYPE (18)
    # ------------------------------------------------------------------
    ("brn_type", "Business Identification Number (BIN)"):          "Business Registration Number",
    ("brn_type", "CAGE"):                                          "Proprietary / Third-party ID",
    ("brn_type", "CAN BUSINESS NUMBER"):                           "Business Registration Number",
    ("brn_type", "CRA Business Number"):                           "Business Registration Number",
    ("brn_type", "CRA Business Number incorporated in NS"):        "Business Registration Number",
    ("brn_type", "CRA Business Number, incorporated in BC"):       "Business Registration Number",
    ("brn_type", "CRA Business Number, incorporated in SK"):       "Business Registration Number",
    ("brn_type", "Canadian Business Number"):                      "Business Registration Number",
    ("brn_type", "Commercial And Government Entity Code"):         "Proprietary / Third-party ID",
    ("brn_type", "DNK ENTITY"):                                    "Business Registration Number",
    ("brn_type", "Federal Business Number"):                       "Business Registration Number",
    ("brn_type", "Financial Conduct Authority Reference Number (GB)"): "Proprietary / Third-party ID",
    ("brn_type", "Provincial Business Number"):                    "Business Registration Number",
    ("brn_type", "SWIFT BIC CODE"):                                "Proprietary / Third-party ID",
    ("brn_type", "Tax ID Number - BN9"):                           "Tax ID Number",
    ("brn_type", "USA FEI NUMBER"):                                "Proprietary / Third-party ID",
    ("brn_type", "USA SAM UEI NUMBER"):                            "Proprietary / Third-party ID",
    ("brn_type", "USA SEC CIK NUMBER"):                            "Proprietary / Third-party ID",
}


# ---------------------------------------------------------------------------
# FieldLookup — the in-memory lookup object for one field.
# ---------------------------------------------------------------------------

@dataclass
class FieldLookup:
    """Pre-flight lookup for a single field.

    Two tiers:
      consistent  — {norm(raw_value): std_value}
                    Used for raw values whose answer is the same regardless
                    of country (or whose conflict was resolved to one answer).
      by_country  — {norm(country): {norm(raw_value): std_value}}
                    Used for raw values whose correct answer depends on country.

    get() checks by_country first, then consistent.
    """
    field_key: str
    consistent: dict[str, str] = dc_field(default_factory=dict)
    by_country: dict[str, dict[str, str]] = dc_field(default_factory=dict)
    _similarity_cache: dict[str, list[tuple[str, str, str, str]]] = dc_field(
        default_factory=dict, init=False, repr=False,
    )
    _normalized_cache: dict[
        str, dict[str, list[tuple[str, str, str, str]]]
    ] = dc_field(default_factory=dict, init=False, repr=False)

    def get(self, raw_value: str, country: str = "") -> Optional[str]:
        """Return the standardized value, or None if not in the lookup."""
        raw_n = _norm(raw_value)
        if not raw_n:
            return None
        country_n = _norm(country)
        # Country-specific first.
        if country_n and country_n in self.by_country:
            result = self.by_country[country_n].get(raw_n)
            if result is not None:
                return result
        # Country-agnostic fallback.
        return self.consistent.get(raw_n)

    def similar(
        self,
        raw_value: str,
        country: str = "",
        *,
        limit: int = 3,
        min_score: float = SIMILARITY_EXAMPLE_MIN_SCORE,
    ) -> list[SimilarityMatch]:
        """Return the closest approved mappings for retrieval-assisted classification.

        Country-specific mappings for the requested country override global
        mappings with the same raw key. Entries from other countries are never
        used as evidence.
        """
        query = _similarity_key(raw_value)
        if not query:
            return []

        country_n = _norm(country)
        candidates = self._similarity_candidates(country_n)

        matches: list[SimilarityMatch] = []
        for candidate_raw, standardized, candidate_country, candidate_key in candidates:
            score = _similarity_score(query, candidate_key)
            if score >= min_score:
                matches.append(SimilarityMatch(
                    raw_value=candidate_raw,
                    standardized_value=standardized,
                    score=score,
                    country=candidate_country,
                ))

        matches.sort(key=lambda m: (-m.score, m.raw_value.casefold()))
        return matches[:limit]

    def _similarity_candidates(
        self, country_n: str,
    ) -> list[tuple[str, str, str, str]]:
        """Build the normalized candidate index once per field/country."""
        cached = self._similarity_cache.get(country_n)
        if cached is not None:
            return cached

        candidates: dict[str, tuple[str, str]] = {
            raw: (std, "") for raw, std in self.consistent.items()
        }
        if country_n and country_n in self.by_country:
            candidates.update({
                raw: (std, country_n)
                for raw, std in self.by_country[country_n].items()
            })
        indexed = [
            (raw, standardized, candidate_country, _similarity_key(raw))
            for raw, (standardized, candidate_country) in candidates.items()
        ]
        self._similarity_cache[country_n] = indexed
        normalized: dict[str, list[tuple[str, str, str, str]]] = defaultdict(list)
        for candidate in indexed:
            normalized[candidate[3]].append(candidate)
        self._normalized_cache[country_n] = dict(normalized)
        return indexed

    def predict_similar(
        self,
        raw_value: str,
        country: str = "",
    ) -> Optional[SimilarityPrediction]:
        """Predict only when a near-identical set of lookup examples agrees.

        Broader semantic/fuzzy matches remain examples for Claude rather than
        becoming automatic decisions.
        """
        query = _similarity_key(raw_value)
        country_n = _norm(country)
        self._similarity_candidates(country_n)
        normalized_matches = self._normalized_cache[country_n].get(query, [])
        if normalized_matches:
            labels = {candidate[1] for candidate in normalized_matches}
            if len(labels) != 1:
                return None
            exact_matches = tuple(
                SimilarityMatch(
                    raw_value=candidate_raw,
                    standardized_value=standardized,
                    score=1.0,
                    country=candidate_country,
                )
                for candidate_raw, standardized, candidate_country, _ in normalized_matches[:3]
            )
            return SimilarityPrediction(
                standardized_value=normalized_matches[0][1],
                score=1.0,
                matches=exact_matches,
            )

        matches = self.similar(raw_value, country, limit=8)
        if not matches:
            return None

        top = matches[0]
        if top.score < SIMILARITY_AUTO_SCORE:
            return None
        if len(query) < 4 and top.score < 1.0:
            return None

        near_top = [m for m in matches if m.score >= max(0.75, top.score - 0.04)]
        if any(m.standardized_value != top.standardized_value for m in near_top):
            return None

        best_other = max(
            (m.score for m in matches if m.standardized_value != top.standardized_value),
            default=0.0,
        )
        if best_other and top.score - best_other < SIMILARITY_AUTO_MARGIN:
            return None

        return SimilarityPrediction(
            standardized_value=top.standardized_value,
            score=top.score,
            matches=tuple(matches[:3]),
        )

    @property
    def consistent_count(self) -> int:
        return len(self.consistent)

    @property
    def country_entry_count(self) -> int:
        return sum(len(v) for v in self.by_country.values())

    @property
    def total_entries(self) -> int:
        return self.consistent_count + self.country_entry_count


# ---------------------------------------------------------------------------
# Internal helpers.
# ---------------------------------------------------------------------------

def _norm(s: str | None) -> str:
    """Whitespace-collapse only — matches _norm_key in standardize_file.py."""
    return " ".join(str(s).split()) if s is not None else ""


def _similarity_key(s: str | None) -> str:
    """Normalize casing, accents, and punctuation for similarity only."""
    if s is None:
        return ""
    decomposed = unicodedata.normalize("NFKD", str(s).casefold())
    without_marks = "".join(
        ch for ch in decomposed if unicodedata.category(ch) != "Mn"
    )
    return " ".join(re.sub(r"[^\w]+", " ", without_marks).split())


def _char_ngrams(text: str, n: int = 3) -> set[str]:
    padded = f" {text} "
    if len(padded) <= n:
        return {padded}
    return {padded[i:i + n] for i in range(len(padded) - n + 1)}


def _similarity_score(left: str, right: str) -> float:
    """Combine spelling, token, and character-pattern similarity."""
    if not left or not right:
        return 0.0
    if left == right:
        return 1.0

    sequence = SequenceMatcher(None, left, right).ratio()
    left_tokens, right_tokens = set(left.split()), set(right.split())
    token_union = left_tokens | right_tokens
    token_score = (
        len(left_tokens & right_tokens) / len(token_union) if token_union else 0.0
    )
    left_grams, right_grams = _char_ngrams(left), _char_ngrams(right)
    gram_score = (
        2 * len(left_grams & right_grams) / (len(left_grams) + len(right_grams))
        if left_grams or right_grams else 0.0
    )
    return max(sequence, token_score, gram_score)


def _fix_label(std_value: str) -> str:
    return _LABEL_FIX.get(std_value, std_value)


def _load_sheet_rows(wb: openpyxl.Workbook, sheet_name: str) -> list[tuple[str, str, str]]:
    """Return (country, raw_value, std_value) triples from one sheet."""
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        country, raw, std = row[0], row[1], row[2]
        if raw is None or std is None:
            continue
        country_n = _norm(country)
        raw_n = _norm(raw)
        std_n = _fix_label(_norm(std))
        if raw_n and std_n:
            rows.append((country_n, raw_n, std_n))
    return rows


def _build_field_lookup(
    field_key: str,
    rows: list[tuple[str, str, str]],
    verbose: bool = True,
) -> FieldLookup:
    """Build a FieldLookup from raw (country, raw_value, std_value) triples."""
    spec = field_registry.get(field_key)
    valid_std_values = set(spec.standard_values)

    # Group: raw_value -> std_value -> set of countries.
    by_raw: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))
    for country, raw, std in rows:
        by_raw[raw][std].add(country)

    lookup = FieldLookup(field_key=field_key)
    skipped_invalid = 0
    skipped_ambiguous = 0
    kept_consistent = 0
    kept_country = 0
    kept_resolved = 0

    for raw, std_map in by_raw.items():
        if len(std_map) == 1:
            # Consistent — one answer regardless of country.
            std_val = next(iter(std_map))
            if std_val not in valid_std_values:
                skipped_invalid += 1
                continue
            lookup.consistent[raw] = std_val
            kept_consistent += 1

        else:
            # Conflicting — check resolution.
            resolution = RESOLUTIONS.get((field_key, raw))

            if resolution is None:
                # No resolution set — skip, let API handle.
                continue

            if resolution == "keep_by_country":
                # Build per-country entries. Skip any country that has
                # more than one answer (within-country ambiguity).
                country_to_std: dict[str, str] = {}
                ambiguous_countries: set[str] = set()

                for std_val, countries in std_map.items():
                    for c in countries:
                        if c in country_to_std:
                            # Same country already mapped — ambiguous.
                            ambiguous_countries.add(c)
                        else:
                            country_to_std[c] = std_val

                for c in ambiguous_countries:
                    country_to_std.pop(c, None)

                skipped_ambiguous += len(ambiguous_countries)

                for c, std_val in country_to_std.items():
                    if std_val not in valid_std_values:
                        skipped_invalid += 1
                        continue
                    lookup.by_country.setdefault(c, {})[raw] = std_val
                    kept_country += 1

            else:
                # Specific resolved value — overrides the master file.
                resolved_std = resolution
                if resolved_std not in valid_std_values:
                    if verbose:
                        print(f"    [!] Resolution {resolved_std!r} for ({field_key}, {raw!r}) "
                              f"is not a valid standard value — skipping.")
                    skipped_invalid += 1
                    continue
                lookup.consistent[raw] = resolved_std
                kept_resolved += 1

    if verbose:
        print(f"    consistent    : {kept_consistent}  "
              f"country-specific: {kept_country}  "
              f"resolved-override: {kept_resolved}  "
              f"skipped(invalid): {skipped_invalid}  "
              f"skipped(within-country-ambiguous): {skipped_ambiguous}")

    return lookup


# ---------------------------------------------------------------------------
# Public API: build and save, load.
# ---------------------------------------------------------------------------

def build_from_master_file(
    master_file: Path = DEFAULT_MASTER_FILE,
    output_file: Path = DEFAULT_LOOKUP_FILE,
    field_keys: list[str] | None = None,
    verbose: bool = True,
) -> dict[str, FieldLookup]:
    """Read the master Excel file and write master_lookup.json.

    Returns the dict of FieldLookup objects (also serialised to output_file).
    """
    if not master_file.exists():
        raise FileNotFoundError(f"Master file not found: {master_file}")

    keys_to_build = field_keys or list(SHEET_BY_FIELD)
    if verbose:
        print(f"Building lookup from: {master_file.name}")
        print(f"Fields: {len(keys_to_build)} of {len(SHEET_BY_FIELD)}")

    wb = openpyxl.load_workbook(master_file, read_only=True, data_only=True)
    lookups: dict[str, FieldLookup] = {}

    for field_key in keys_to_build:
        sheet_name = SHEET_BY_FIELD[field_key]
        spec = field_registry.get(field_key)
        if verbose:
            print(f"\n  {spec.display_name}")
        try:
            rows = _load_sheet_rows(wb, sheet_name)
        except KeyError as exc:
            if verbose:
                print(f"    SKIPPED — {exc}")
            continue

        if verbose:
            print(f"    rows loaded   : {len(rows)}")

        lookup = _build_field_lookup(field_key, rows, verbose=verbose)
        lookups[field_key] = lookup

        if verbose:
            print(f"    total entries : {lookup.total_entries} "
                  f"({lookup.consistent_count} consistent, "
                  f"{lookup.country_entry_count} country-specific)")

    wb.close()

    # Serialise to JSON.
    # Structure: {field_key: {"consistent": {...}, "by_country": {country: {...}}}}
    serializable = {
        fk: {
            "consistent": lk.consistent,
            "by_country": lk.by_country,
        }
        for fk, lk in lookups.items()
    }
    output_file.parent.mkdir(parents=True, exist_ok=True)
    output_file.write_text(json.dumps(serializable, indent=2, ensure_ascii=False), encoding="utf-8")

    if verbose:
        total = sum(lk.total_entries for lk in lookups.values())
        print(f"\nTotal lookup entries : {total}")
        print(f"Saved to             : {output_file}")

    return lookups


def _s3_client(cfg: dict):
    """Return a boto3 S3 client from a config dict."""
    import boto3
    return boto3.client(
        "s3",
        region_name=cfg["region"],
        aws_access_key_id=cfg["access_key"],
        aws_secret_access_key=cfg["secret_key"],
    )


def _load_json_from_s3(cfg: dict) -> dict:
    """Download and parse master_lookup.json from S3. Returns {} on any error."""
    try:
        client = _s3_client(cfg)
        response = client.get_object(Bucket=cfg["bucket"], Key=cfg["key"])
        return json.loads(response["Body"].read().decode("utf-8"))
    except Exception as exc:
        print(f"[S3] Could not load lookup from S3: {exc} — falling back to local file.")
        return {}


def _save_json_to_s3(cfg: dict, data: dict) -> bool:
    """Upload data as master_lookup.json to S3. Returns True on success."""
    try:
        client = _s3_client(cfg)
        client.put_object(
            Bucket=cfg["bucket"],
            Key=cfg["key"],
            Body=json.dumps(data, indent=2, ensure_ascii=False).encode("utf-8"),
            ContentType="application/json",
        )
        return True
    except Exception as exc:
        print(f"[S3] Could not save lookup to S3: {exc}")
        return False


def merge_api_results(
    field_key: str,
    new_mappings: dict,
    country_dependent: bool,
    lookup_file: Path = DEFAULT_LOOKUP_FILE,
) -> int:
    """Merge new mappings into the lookup (S3 if configured, else local file).

    Only adds entries not already present — never overwrites existing mappings.
    Returns the count of new entries added.
    """
    if not new_mappings:
        return 0

    import config as _config
    s3_cfg = _config.get_s3_config()

    # Load existing data from S3 or local file.
    if s3_cfg:
        data = _load_json_from_s3(s3_cfg)
        if not data and lookup_file.exists():
            # S3 empty/new — seed from local file.
            data = json.loads(lookup_file.read_text(encoding="utf-8"))
    elif lookup_file.exists():
        data = json.loads(lookup_file.read_text(encoding="utf-8"))
    else:
        return 0

    field_data = data.get(field_key, {"consistent": {}, "by_country": {}})
    consistent = field_data.get("consistent", {})
    by_country = field_data.get("by_country", {})

    added = 0
    for (country, raw_val), std_val in new_mappings.items():
        # Normalise keys (collapse whitespace) to match how the pipeline looks them up.
        raw_val = " ".join(str(raw_val).split()) if raw_val else ""
        country = " ".join(str(country).split()) if country else ""
        if not raw_val or not std_val:
            continue
        if country_dependent and country:
            country_map = by_country.setdefault(country, {})
            if raw_val not in country_map and raw_val not in consistent:
                country_map[raw_val] = std_val
                added += 1
        else:
            if raw_val not in consistent:
                consistent[raw_val] = std_val
                added += 1

    field_data["consistent"] = consistent
    field_data["by_country"] = by_country
    data[field_key] = field_data

    # Always save when called — even if added==0, the caller may be forcing a refresh.
    if s3_cfg:
        _save_json_to_s3(s3_cfg, data)
    lookup_file.parent.mkdir(parents=True, exist_ok=True)
    lookup_file.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    return added


def load_lookup(lookup_file: Path = DEFAULT_LOOKUP_FILE) -> dict[str, FieldLookup]:
    """Load the lookup from S3 if configured, otherwise from the local file.

    S3 is the source of truth when credentials are present — it always has the
    latest version regardless of which machine last updated it.
    Falls back to local file gracefully so local development keeps working.
    """
    import config as _config
    s3_cfg = _config.get_s3_config()

    if s3_cfg:
        data = _load_json_from_s3(s3_cfg)
        if data:
            # Also write a local copy so the app works if S3 is temporarily unreachable.
            lookup_file.parent.mkdir(parents=True, exist_ok=True)
            lookup_file.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
        else:
            # S3 failed or empty — fall back to local.
            if lookup_file.exists():
                data = json.loads(lookup_file.read_text(encoding="utf-8"))
    elif lookup_file.exists():
        data = json.loads(lookup_file.read_text(encoding="utf-8"))
    else:
        return {}

    lookups: dict[str, FieldLookup] = {}
    for fk, fdata in data.items():
        lk = FieldLookup(field_key=fk)
        lk.consistent = fdata.get("consistent", {})
        lk.by_country = fdata.get("by_country", {})
        lookups[fk] = lk
    return lookups
