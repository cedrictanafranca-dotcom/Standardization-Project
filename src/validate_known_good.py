r"""Step 7 — Validate against known-good data.

Runs the SAME classification pipeline the app uses (src/classifier.py,
field-aware via src/fields.py) against real prior classifications, and reports
accuracy + mismatches. This is the accuracy check the brief calls for before
trusting the tool on new data (build plan Step 7).

Ground truth comes from the master mapping file already referenced in the
brief (Section 9 / Section 12): "[FINAL]- ALL Countries GG Standardization
Remapping.xlsx", one sheet per field, columns (Country, raw Input/Value,
Standardized Value). Sheet names don't match field keys cleanly (the brief
warns of this), so this script hardcodes the sheet-name mapping below.

Important nuance this script makes explicit rather than hiding: the current
pipeline (Step 3/5) deduplicates and classifies raw values WITHOUT their
country — it sends one flat list of unique raw values with no per-row country
context. For most raw values that's fine (one raw string always means one
standard value everywhere), but the ground-truth data shows a real minority of
raw values are genuinely COUNTRY-DEPENDENT (e.g. "President" is Board Member
in the UK/GI/IE/MT but Executive Management elsewhere, per the prompt's own
regional rule). Those can't be fairly scored as right/wrong against a single
expected answer with the pipeline as it stands today, so this script splits
ground truth into:
  - "consistent"       — one correct answer regardless of country (scored)
  - "country-dependent" — 2+ different correct answers depending on country
                          (reported separately, NOT scored as pass/fail)

Run:
    .venv\Scripts\python.exe src\validate_known_good.py --field positions_designations
    .venv\Scripts\python.exe src\validate_known_good.py --field business_legal_form
    .venv\Scripts\python.exe src\validate_known_good.py --field business_status --live
"""

from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:  # pragma: no cover
    pass

import openpyxl

import config
import fields
from classifier import MockClaudeClient, RealClaudeClient, classify_values

DEFAULT_MASTER_FILE = (
    Path.home() / "Downloads" / "[FINAL]- ALL Countries GG Standardization Remapping.xlsx"
)

# Sheet names don't match field keys cleanly (brief Section 9) — mapped by hand.
SHEET_BY_FIELD = {
    "business_legal_form": "StandardizedIncorporationDetail",
    "positions_designations": "Universal Position + Designatio",
    "business_status": "Sheet20",
    "directors_officers_status": "Status - DirectorsOfficers",
    "ownership_relationship_status": "Status - OwnershipRelationship",
    "psc_beneficiary_type": "UniversalBeneficiaryType",
    "directors_officers_type": "Type - DirectorsOfficers",
    "ownership_relationship_type": "Type - Ownership Relationship T",
    "business_entity_type": "BusinessEntityType",
    "brn_type": "BRN Type",
}

# Master file predates the brief's "Other / Unknown" -> "Other / Unclassified"
# normalization (Section 11) — normalize ground truth the same way so the
# comparison is apples-to-apples, not penalizing the tool for a label fix the
# brief explicitly called for.
_LABEL_FIX = {"Other / Unknown": "Other / Unclassified", "Unknown": "Other / Unclassified"}


def _norm(s) -> str:
    return " ".join(str(s).split()) if s is not None else ""


def load_ground_truth(master_file: Path, field_key: str) -> dict[str, set[str]]:
    """Return {normalized_raw_value: {set of standardized values seen}}."""
    sheet_name = SHEET_BY_FIELD[field_key]
    wb = openpyxl.load_workbook(master_file, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    by_raw: dict[str, set[str]] = defaultdict(set)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        raw, std = row[1], row[2]
        if raw is None or std is None:
            continue
        raw_n = _norm(raw)
        std_n = _norm(std)
        std_n = _LABEL_FIX.get(std_n, std_n)
        if raw_n:
            by_raw[raw_n].add(std_n)
    return by_raw


def split_consistent_vs_ambiguous(
    by_raw: dict[str, set[str]],
) -> tuple[dict[str, str], dict[str, set[str]]]:
    consistent = {k: next(iter(v)) for k, v in by_raw.items() if len(v) == 1}
    ambiguous = {k: v for k, v in by_raw.items() if len(v) > 1}
    return consistent, ambiguous


def main() -> int:
    parser = argparse.ArgumentParser(description="Step 7 — validate against known-good data")
    parser.add_argument("--field", required=True, choices=sorted(fields.FIELDS))
    parser.add_argument("--master-file", type=Path, default=DEFAULT_MASTER_FILE)
    parser.add_argument("--batch-size", type=int, default=100)
    parser.add_argument("--live", action="store_true", help="use the real Anthropic API")
    parser.add_argument(
        "--limit", type=int, default=None,
        help="cap how many consistent raw values get validated (useful with --live "
             "to control API cost); default: no cap",
    )
    parser.add_argument(
        "--show-mismatches", type=int, default=25,
        help="max mismatch rows to print (all are counted regardless)",
    )
    args = parser.parse_args()

    if not args.master_file.exists():
        print(f"Master mapping file not found: {args.master_file}")
        return 1

    spec = fields.get(args.field)
    print(f"Field       : {spec.display_name}  ({args.field})")
    print(f"Sheet       : {SHEET_BY_FIELD[args.field]!r} in {args.master_file.name}")

    by_raw = load_ground_truth(args.master_file, args.field)
    consistent, ambiguous = split_consistent_vs_ambiguous(by_raw)

    print(f"Ground truth: {len(by_raw)} unique raw values "
          f"({len(consistent)} consistent, {len(ambiguous)} country-dependent)")

    if ambiguous:
        print(f"\n  {len(ambiguous)} raw values have MORE THAN ONE correct answer "
              "depending on country — the current pipeline classifies raw values "
              "without country context, so these can't be fairly scored right now.")
        print("  Excluded from the accuracy score below. Examples:")
        for k in list(ambiguous)[:5]:
            print(f"      {k!r} -> {sorted(ambiguous[k])}")

    raw_values = list(consistent)
    total_available = len(raw_values)
    if args.limit and len(raw_values) > args.limit:
        raw_values = raw_values[: args.limit]
        print(f"\n  NOTE: capped to {args.limit} of {total_available} consistent values "
              f"(--limit). {total_available - args.limit} were NOT validated this run.")

    system_prompt = spec.load_prompt()
    client = RealClaudeClient() if args.live else MockClaudeClient(field_key=args.field)
    print(f"Client      : {client.name}")
    if not args.live and args.field not in ("positions_designations", "business_legal_form"):
        print(f"  NOTE: mock has no keyword heuristic for {args.field!r} — expect near-zero "
              "accuracy here. This validates the SCRIPT/comparison logic, not real accuracy; "
              "use --live for a real accuracy check on this field.")

    # Batch through the same classify_values() the app uses.
    mismatches: list[tuple[str, str, str]] = []
    correct = 0
    # Step 8 check: does confidence actually correlate with correctness?
    # {confidence_level: [n_correct, n_total]}
    by_confidence: dict[str, list[int]] = defaultdict(lambda: [0, 0])
    for i in range(0, len(raw_values), args.batch_size):
        chunk = raw_values[i : i + args.batch_size]
        batch = classify_values(chunk, system_prompt, client)
        for r in batch.results:
            expected = consistent[r.raw_value]
            conf_key = r.confidence or "(missing)"
            is_correct = r.standardized_value == expected
            by_confidence[conf_key][1] += 1
            if is_correct:
                correct += 1
                by_confidence[conf_key][0] += 1
            else:
                mismatches.append((r.raw_value, expected, r.standardized_value))

    scored = len(raw_values)
    accuracy = (correct / scored * 100) if scored else 0.0

    print("\n" + "-" * 68)
    print(f"  scored            : {scored} raw values (consistent-answer subset)")
    print(f"  correct           : {correct}")
    print(f"  mismatches        : {len(mismatches)}")
    print(f"  accuracy          : {accuracy:.1f}%")
    print("-" * 68)

    print("\nAccuracy by confidence level (does confidence predict correctness?):")
    print(f"  {'LEVEL':<10} {'CORRECT/TOTAL':<16} ACCURACY")
    for level in ("HIGH", "MEDIUM", "LOW", "(missing)"):
        if level not in by_confidence:
            continue
        n_correct, n_total = by_confidence[level]
        pct = (n_correct / n_total * 100) if n_total else 0.0
        print(f"  {level:<10} {f'{n_correct}/{n_total}':<16} {pct:.1f}%")

    if mismatches:
        print(f"\nMismatches (showing up to {args.show_mismatches} of {len(mismatches)}):")
        print(f"  {'RAW VALUE':<35} {'EXPECTED':<30} GOT")
        for raw, expected, got in mismatches[: args.show_mismatches]:
            print(f"  {raw:<35.35} {expected:<30.30} {got}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
