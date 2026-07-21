r"""Audit conflicting mappings in the master remapping file.

A "conflict" is a raw value that maps to MORE THAN ONE standardized value
across different rows in the master file (i.e. the same raw string gets
different answers depending on the country).

These conflicts were deferred in Step 7 — this script enumerates them across
all 10 fields, explains WHY they conflict (which countries get which answer),
and exports a structured Excel report for human review before the master lookup
table is seeded.

Output (in output/ directory):
    conflict_audit.xlsx  — one row per (field, raw_value, country, std_value);
                           conflicts are flagged so they can be filtered in Excel.
    conflict_summary.txt — plain-text summary, one section per field.

Run:
    .venv\Scripts\python.exe src\audit_conflicts.py
    .venv\Scripts\python.exe src\audit_conflicts.py --master-file "path\to\file.xlsx"
    .venv\Scripts\python.exe src\audit_conflicts.py --fields positions_designations business_legal_form
"""

from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

import openpyxl
import pandas as pd

import config
import fields

DEFAULT_MASTER_FILE = (
    Path.home() / "Downloads" / "[FINAL]- ALL Countries GG Standardization Remapping.xlsx"
)

# Sheet names in the master file — must be mapped by hand (brief Section 9).
SHEET_BY_FIELD = {
    "business_legal_form":         "StandardizedIncorporationDetail",
    "positions_designations":      "Universal Position + Designatio",
    "business_status":             "Sheet20",
    "directors_officers_status":   "Status - DirectorsOfficers",
    "ownership_relationship_status": "Status - OwnershipRelationship",
    "psc_beneficiary_type":        "UniversalBeneficiaryType",
    "directors_officers_type":     "Type - DirectorsOfficers",
    "ownership_relationship_type": "Type - Ownership Relationship T",
    "business_entity_type":        "BusinessEntityType",
    "brn_type":                    "BRN Type",
}

# Same label fix applied in validate_known_good.py — compare apples-to-apples.
_LABEL_FIX = {"Other / Unknown": "Other / Unclassified", "Unknown": "Other / Unclassified"}


def _norm(s) -> str:
    return " ".join(str(s).split()) if s is not None else ""


def load_field_rows(master_file: Path, field_key: str) -> list[tuple[str, str, str]]:
    """Return list of (country, raw_value, standardized_value) for this field's sheet."""
    sheet_name = SHEET_BY_FIELD[field_key]
    wb = openpyxl.load_workbook(master_file, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        country, raw, std = row[0], row[1], row[2]
        if raw is None or std is None:
            continue
        country_n = _norm(country)
        raw_n = _norm(raw)
        std_n = _norm(std)
        std_n = _LABEL_FIX.get(std_n, std_n)
        if raw_n:
            rows.append((country_n, raw_n, std_n))
    wb.close()
    return rows


def find_conflicts(
    rows: list[tuple[str, str, str]],
) -> dict[str, dict[str, list[str]]]:
    """Return {raw_value: {std_value: [countries]}} for raw values with >1 std_value."""
    # Build {raw_value: {std_value: set(countries)}}
    by_raw: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))
    for country, raw, std in rows:
        by_raw[raw][std].add(country)

    # Keep only raw values with more than one distinct standardized value.
    conflicts: dict[str, dict[str, list[str]]] = {}
    for raw, std_map in by_raw.items():
        if len(std_map) > 1:
            conflicts[raw] = {std: sorted(countries) for std, countries in std_map.items()}
    return conflicts


def audit_all_fields(
    master_file: Path,
    field_keys: list[str],
) -> tuple[list[dict], dict[str, int]]:
    """
    Returns:
        records  — flat list of dicts, one per (field, raw_value, std_value, country) row,
                   with a 'is_conflict' bool and a 'conflict_group' int per raw value.
        summary  — {field_key: conflict_count}
    """
    records: list[dict] = []
    summary: dict[str, int] = {}
    conflict_group = 0

    for field_key in field_keys:
        spec = fields.get(field_key)
        print(f"  Loading {spec.display_name} ...", end=" ", flush=True)
        try:
            rows = load_field_rows(master_file, field_key)
        except KeyError as exc:
            print(f"SKIPPED — {exc}")
            summary[field_key] = 0
            continue

        conflicts = find_conflicts(rows)
        summary[field_key] = len(conflicts)
        print(f"{len(rows)} rows, {len(conflicts)} conflicting raw values")

        # Build lookup: is this raw value a conflict?
        conflict_raws = set(conflicts)

        # Emit one record per (raw_value, std_value, country) triple,
        # with conflict metadata attached.
        # Build full by_raw (all values, not just conflicts) for consistent output.
        by_raw: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))
        for country, raw, std in rows:
            by_raw[raw][std].add(country)

        # Only emit conflict rows (consistent rows go to the lookup, not the audit).
        for raw in sorted(conflicts):
            conflict_group += 1
            group_id = conflict_group
            std_map = conflicts[raw]
            for std_val, countries in sorted(std_map.items()):
                for country in countries:
                    records.append({
                        "Field": spec.display_name,
                        "Field Key": field_key,
                        "Conflict Group": group_id,
                        "Raw Value": raw,
                        "Standardized Value": std_val,
                        "Country": country,
                        "Num Answers": len(std_map),
                        "All Answers": " | ".join(sorted(std_map)),
                    })

    return records, summary


def print_summary(summary: dict[str, int], total_records: int) -> None:
    print("\n" + "=" * 68)
    print("CONFLICT AUDIT SUMMARY")
    print("=" * 68)
    total_conflicts = sum(summary.values())
    print(f"  Total conflicting raw values : {total_conflicts}")
    print(f"  Total rows in audit file     : {total_records}")
    print()
    print(f"  {'Field':<40} Conflicts")
    print(f"  {'-'*40} ---------")
    for field_key, count in summary.items():
        spec = fields.get(field_key)
        flag = " <-- has conflicts" if count else ""
        print(f"  {spec.display_name:<40} {count}{flag}")
    print("=" * 68)
    print()
    print("What these conflicts mean:")
    print("  Each conflicting raw value appears in the master file with 2+ different")
    print("  standardized values, depending on the country. They are NOT data errors")
    print("  by default — many are genuine country-specific rules (e.g. 'President'")
    print("  -> Board Member in UK/GI/IE/MT, Executive Management elsewhere).")
    print()
    print("What to do with this file:")
    print("  Open conflict_audit.xlsx. For each Conflict Group:")
    print("    A) GENUINE country rule  -> mark 'Resolution: keep_by_country'")
    print("       The lookup table will use (country, raw_value) as the key.")
    print("    B) DATA QUALITY issue    -> mark 'Resolution: <correct_std_value>'")
    print("       The lookup table will use raw_value only, with your chosen value.")
    print("  Unresolved entries will be sent to the Claude API as normal.")


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Audit conflicting (multi-answer) raw values in the master mapping file"
    )
    parser.add_argument(
        "--master-file", type=Path, default=DEFAULT_MASTER_FILE,
        help=f"path to master remapping xlsx (default: {DEFAULT_MASTER_FILE})",
    )
    parser.add_argument(
        "--fields", nargs="+", default=None,
        choices=sorted(SHEET_BY_FIELD),
        metavar="FIELD",
        help="subset of fields to audit (default: all 10)",
    )
    parser.add_argument(
        "--output-dir", type=Path, default=config.OUTPUT_DIR,
    )
    args = parser.parse_args()

    if not args.master_file.exists():
        print(f"[!] Master mapping file not found: {args.master_file}")
        print("    Check the path or pass --master-file.")
        return 1

    field_keys = args.fields or list(SHEET_BY_FIELD)
    print(f"Master file : {args.master_file.name}")
    print(f"Fields      : {len(field_keys)} of 10")
    print()

    records, summary = audit_all_fields(args.master_file, field_keys)

    if not records:
        print("\nNo conflicts found across the selected fields.")
        return 0

    # -- Excel report -------------------------------------------------------
    df = pd.DataFrame(records)

    # Add a blank "Resolution" column for the reviewer to fill in.
    df["Resolution"] = ""

    # Reorder for readability.
    col_order = [
        "Field", "Conflict Group", "Raw Value", "Num Answers", "All Answers",
        "Standardized Value", "Country", "Resolution", "Field Key",
    ]
    df = df[[c for c in col_order if c in df.columns]]

    out_xlsx = args.output_dir / "conflict_audit.xlsx"
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Conflicts")
        ws = writer.sheets["Conflicts"]

        # Freeze top row and set column widths.
        ws.freeze_panes = "A2"
        widths = {
            "A": 30,  # Field
            "B": 14,  # Conflict Group
            "C": 40,  # Raw Value
            "D": 12,  # Num Answers
            "E": 50,  # All Answers
            "F": 30,  # Standardized Value
            "G": 30,  # Country
            "H": 25,  # Resolution
            "I": 25,  # Field Key
        }
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width

    print_summary(summary, len(records))
    print(f"Audit saved : {out_xlsx}")

    # -- Plain-text summary -------------------------------------------------
    out_txt = args.output_dir / "conflict_summary.txt"
    lines = []
    for field_key, count in summary.items():
        if count == 0:
            continue
        spec = fields.get(field_key)
        lines.append(f"\n{'='*60}")
        lines.append(f"FIELD: {spec.display_name}  ({count} conflicting raw values)")
        lines.append(f"{'='*60}")
        field_rows = [r for r in records if r["Field Key"] == field_key]
        # Group by raw value.
        by_raw: dict[str, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))
        for r in field_rows:
            by_raw[r["Raw Value"]][r["Standardized Value"]].append(r["Country"])
        for raw, std_map in sorted(by_raw.items()):
            lines.append(f"\n  Raw: {raw!r}")
            for std_val, countries in sorted(std_map.items()):
                lines.append(f"    -> {std_val!r}  [{', '.join(countries[:8])}{'...' if len(countries) > 8 else ''}]")

    out_txt.write_text("\n".join(lines), encoding="utf-8")
    print(f"Summary txt : {out_txt}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
