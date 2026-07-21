r"""Apply conflict resolutions to conflict_audit.xlsx.

Reads output/conflict_audit.xlsx (produced by audit_conflicts.py), fills in
the 'Resolution' column from RESOLUTIONS (defined in master_lookup.py — the
authoritative source), and saves the file.

Run:
    .venv\Scripts\python.exe src\apply_resolutions.py
"""

from __future__ import annotations

import sys
from collections import Counter

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

import openpyxl

import config
from master_lookup import RESOLUTIONS  # authoritative source — edit decisions there

AUDIT_FILE = config.OUTPUT_DIR / "conflict_audit.xlsx"


def main() -> int:
    if not AUDIT_FILE.exists():
        print(f"[!] Audit file not found: {AUDIT_FILE}")
        print("    Run audit_conflicts.py first.")
        return 1

    wb = openpyxl.load_workbook(AUDIT_FILE)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    try:
        col_field_key  = headers.index("Field Key") + 1
        col_raw_value  = headers.index("Raw Value") + 1
        col_resolution = headers.index("Resolution") + 1
    except ValueError as exc:
        print(f"[!] Missing expected column in audit file: {exc}")
        return 1

    applied = 0
    unmatched_keys: set[tuple[str, str]] = set()

    for row in ws.iter_rows(min_row=2):
        field_key = row[col_field_key - 1].value
        raw_value = row[col_raw_value - 1].value
        if field_key is None or raw_value is None:
            continue
        key = (str(field_key).strip(), str(raw_value).strip())
        if key in RESOLUTIONS:
            row[col_resolution - 1].value = RESOLUTIONS[key]
            applied += 1
        else:
            unmatched_keys.add(key)

    wb.save(AUDIT_FILE)

    print(f"Applied {applied} resolution(s) to {AUDIT_FILE.name}")
    if unmatched_keys:
        print(f"\n  WARNING: {len(unmatched_keys)} row(s) had no matching resolution key:")
        for k in sorted(unmatched_keys):
            print(f"    {k}")
    else:
        print("  All rows matched — no gaps.")

    by_resolution: Counter = Counter(RESOLUTIONS.values())
    print(f"\nResolution breakdown across {len(RESOLUTIONS)} raw values:")
    for res, count in sorted(by_resolution.items(), key=lambda x: -x[1]):
        print(f"  {count:3d}  {res}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
